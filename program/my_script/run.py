"""
네이버 블로그 도구
"""

# 시스템
import sys, os, time, math
from multiprocessing import Process, Queue as MPQueue

# 크롬 브라우저 제어
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException

# 안티캡챠
from python3_anticaptcha import ImageToTextTask 

# 엑셀
from openpyxl import load_workbook  

# 데스크탑 앱 모듈
from PyQt5.QtCore import QObject, pyqtSignal, pyqtSlot, QThread
from PyQt5.QtWidgets import (QApplication, QMainWindow, QFileDialog, 
    QDesktopWidget, QPushButton, QMessageBox)

# 마우스/키보드 컨트롤
import pyautogui 

# 클립보드
import pyperclip 

# 나이트관련 키워드 # 블로그 제목 검색용
from my_keyword import * 


# 멀티프로세스 메시지 큐 종료 플래그
JOBS_DONE = 'jobs_done' 


"""
텍스트 안에 키워드가 하나라도 있는지 확인

text: str
keywords: str list
return: bool
"""
def has_keyword(text, keywords):
    return any(keyword in text for keyword in keywords)


"""
웹 원소를 새탭으로 열기

browser: webdriver
element: webelement
no return
"""
def open_new_tab(browser, element):
    ActionChains(browser) \
    .key_down(Keys.CONTROL) \
    .click(element) \
    .key_up(Keys.CONTROL) \
    .perform()


"""
웹 원소가 있던 없던 클릭 시도

browser: webdriver
xpath: str
timeout: int
no return
"""
def click_maybe_or_not(browser, xpath, timeout=8):
    try:
        elem = WebDriverWait(browser, timeout=timeout).until(
            EC.presence_of_element_located((By.XPATH, xpath))) 

    except:
        pass

    try:
        browser.execute_script('arguments[0].click();', elem) 

    except:
        pass    


"""
웹 원소가 있는지 확인

browser: webdriver
xpath: str
timeout: int
return bool webelement tuple
"""
def find_element_wait_for(browser, xpath, timeout=3):
    try:
        elem = WebDriverWait(browser, timeout=timeout) \
            .until(EC.presence_of_element_located((By.XPATH, xpath)))

    except TimeoutException:
        return False, None

    return True, elem


"""
클릭 스크립트

browser: webdriver
element: webelement
"""
def click(browser, element):
    browser.execute_script('arguments[0].click();', element)


"""
웹 원소를 기다렸다가 클릭

browser: webdriver
xpath: str
timeout: int
return: bool
"""
def wait_and_click(browser, xpath, timeout=3):
    succ, elem = find_element_wait_for(browser, xpath)

    if succ:
        click(browser, elem)
        
    return succ


"""
크롬브라우저와 엑셀파일 첫 행 지우고 반환하기

excel_filename: str
read_only: bool
return webdriver, cell list list
"""
def open_browser_and_excel_file(excel_filename, read_only=True):
    book = load_workbook(excel_filename, read_only=read_only)
    
    # 첫번째 시트
    sheet = book.worksheets[0] 

    """
    시트 포맷
    1 [id][pw][subject][content1][content2][content3][tags][imgs][year][month][day][hour][minute][is_open]
    2 [my_id][my_pw][my_subject]...
    3 [other_id][other_pw][other_subject]...
    4 ...
    """
    # 첫번째 행을 제외한 나머지 행
    rows = sheet.iter_rows(min_row=2) 

    # 크롬 시크릿 창으로 열기 설정
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--incognito')

    browser = webdriver.Chrome(executable_path='./program/chromedriver.exe', chrome_options=chrome_options)
    return browser, rows


"""
에디터 창에서 이미지 올리기
"""
def upload_image(element_image, image_paths):

    # 현재 디렉터리
    cwd = os.getcwd() + '/'

    # 이미지가 여러장일 경우 분리
    image_paths = [image_path for image_path in image_paths.splitlines() if image_path is not ''] 

    # 이미지의 경로와 파일이름 분리
    path_filenames = [[]]
    path_filenames.pop()

    for image_path in image_paths:
        index = image_path.rfind('/')

        if index == -1:
            path_filenames.append([None, image_path])

        else:
            path_filenames.append([image_path[:index+1], image_path[index+1:]])

    # 이미지 업로드
    for pf in path_filenames:

        # 이미지 업로드 버튼 클릭
        element_image.click() 
        time.sleep(2)

        # 탭 5회 누르기
        for _ in range(5):
            pyautogui.press('tab')

        # 엔터 누르기
        pyautogui.press('enter')

        # 경로 입력
        if pf[0] != None:
            pyautogui.typewrite(cwd + pf[0])

        else:
            pyautogui.typewrite(cwd)

        # 엔터 누르기
        pyautogui.press('enter')

        # 탭 5회 누르기
        for _ in range(5):
            pyautogui.press('tab')

        # 파일 입력
        pyautogui.typewrite(pf[1])

        # 엔터 누르기
        pyautogui.press('enter')
        time.sleep(5)    


"""
데스크탑 앱
"""
class MainWindow(QMainWindow):

    def __init__(self):
        super().__init__()

        # 앱 이름
        self.setWindowTitle('네이버 블로그 도구')         

        # 앱 크기
        self.setGeometry(0, 0, 350, 190) 

        # 앱을 화면 가운데로 옮기기
        rect = self.frameGeometry()
        center = QDesktopWidget().availableGeometry().center()
        rect.moveCenter(center)
        self.move(rect.topLeft())

        # 새글 쓰기 버튼 생성
        self.create_button('새글 쓰기', 150, 50, 10, 10, self.write_new_post) 

        # 마지막 글 지우기 버튼 생성
        self.create_button('마지막 글 지우기', 150, 50, 10, 70, self.delete_last_post) 

        # 다른 글 바꾸기 버튼 생성
        self.create_button('다른 글 바꾸기', 150, 50, 10, 130, self.modify_other_post) 

        # 멀티프로세스용 메시지 큐
        self.message_queue = MPQueue() 

        # 팝업 메시지박스 스레드
        self.ui_thread = UIThread(self.message_queue)
        self.ui_thread.popup_message.connect(self.create_message_box)
        self.ui_thread.start()


    """
    팝업 메시지
    """
    @pyqtSlot(str)
    def create_message_box(self, message):
        QMessageBox.about(self, '', message)


    """
    버튼 만들기
    """
    def create_button(self, text, width, height, x, y, event):
        btn = QPushButton(text, self) 
        
        # 크기
        btn.resize(width, height) 
        
        # 위치 (좌상단 기준)
        btn.move(x, y) 
        
        # 클릭 이벤트
        btn.clicked.connect(event) 


    """
    엑셀 파일 열기
    """
    def open_excel_file(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog

        filename, _ = QFileDialog \
            .getOpenFileName(self, '엑셀 파일 열기', './', '엑셀 파일 (*.xlsx)', options=options) 

        success = True if filename else False
        return (success, filename)

    
    """
    새글 쓰기
    """
    def write_new_post(self):
        success, filename = self.open_excel_file()

        # 엑셀 파일 열기 실패시 종료
        if not success:
            return

        # 새글 쓰기 프로세스 시작
        writer_proc = Process(target=writer_process, args=(filename, self.message_queue))
        writer_proc.start()
        

    """
    마지막 글 지우기
    """
    def delete_last_post(self):
        success, filename = self.open_excel_file()

        # 엑셀 파일 열기 실패시 종료
        if not success:
            return

        # 마지막 글 지우기 프로세스 시작
        deleter_proc = Process(target=deleter_process, args=(filename, self.message_queue))
        deleter_proc.start()


    """
    다른 글 바꾸기
    """
    def modify_other_post(self):
        success, filename = self.open_excel_file()

        # 엑셀 파일 열기 실패시 종료
        if not success:
            return

        # 다른 글 바꾸기 프로세스 시작
        modifier_proc = Process(target=modifier_process, args=(filename, self.message_queue))
        modifier_proc.start()
        

"""
다른 프로세스로부터 받은 메시지 처리 스레드
"""
class UIThread(QThread):

    # 팝업 메시지 시그널 생성
    popup_message = pyqtSignal(str) 

    def __init__(self, message_queue):
        super().__init__()
        self.message_queue = message_queue


    """
    메시지 처리 루프
    """
    def run(self):
        while True:
            msg = self.message_queue.get()

            if msg == JOBS_DONE:
                break

            # 메시지 팝업
            self.popup_message.emit(msg) 


"""
공통 로그인 프로세스 작업
"""
def common_login_process_work(excel_filename, message_queue, work):
    browser, rows = open_browser_and_excel_file(excel_filename)

    # 크롬 열기 실패시 종료
    if not browser:
        message_queue.put('크롬을 열 수 없습니다.')
        return

    # 이전 행 아이디
    prev_id = '!@#$' 
    for row in rows:
        my_id = row[0].value

        # 아이디가 비어있으면 종료
        if (my_id is None) or (my_id is ''): 
            break

        my_pw = row[1].value
        subject = row[2].value
        content = row[3].value
        content2 = row[4].value
        content3 = row[5].value
        tags = row[6].value
        imgs = row[7].value
        year = row[8].value
        month = row[9].value
        day = row[10].value
        hour = row[11].value
        minute = row[12].value
        is_open = row[13].value

        # 아이디가 다르면 새로 로그인하기
        if prev_id != my_id:
            prev_id = my_id

            # 로그아웃 한번 하기
            succ = naver_logout(browser, message_queue) 

            if not succ:
                return

            print('로그인 시도')

            # 로그인 하기
            succ = naver_login(browser, message_queue, my_id, my_pw) 

            if not succ:
                return
            
            print('로그인 성공')

        work(browser, message_queue, my_id, subject, content, content2, content3, tags, imgs, year, month, day, hour, minute, is_open)
        

"""
새글 쓰기 프로세스
"""
def writer_process(excel_filename, message_queue):
    common_login_process_work(excel_filename, message_queue, write_new_post)
    print('새글 쓰기 프로세스 종료')


"""
네이버 로그아웃하기
"""
def naver_logout(browser, message_queue):

    # 네이버 로그아웃 
    browser.get('https://nid.naver.com/nidlogin.logout') 

    succ, _ = find_element_wait_for(browser, '//*[@id="content"]/div[1]/p')

    if not succ:
        message_queue.put('타임아웃: 네이버 로그아웃')
        return False

    return True


"""
네이버 로그인하기
"""
def naver_login(browser, message_queue, my_id, my_pw):

    # 네이버 로그인 창 열기
    browser.get('https://nid.naver.com/nidlogin.login') 

    # 아이디 입력 필드
    succ, elem_id = find_element_wait_for(browser, '//*[@id="id"]')

    if not succ:
        message_queue.put('타임아웃: 네이버 로그인')
        return False

    # 아이디 쓰기
    elem_id.send_keys(my_id) 

    # 비밀번호 쓰기
    elem_pw = browser.find_element_by_xpath('//*[@id="pw"]')    
    elem_pw.send_keys(my_pw)
    elem_pw.submit()

    # 네이버 메인
    succ, _ = find_element_wait_for(browser, '//*[@id="PM_ID_ct"]')

    if succ:
        return True

    else:
        print('로그인 실패')

        # 캡챠 창
        succ, _ = find_element_wait_for(browser, '//*[@id="chptcha"]')        

        if succ:
            return anti_captcha(browser, message_queue, my_id, my_pw)

        else:
            print('로그인 재시도')
            return naver_login(browser, message_queue, my_id, my_pw)            


"""
안티캡챠
"""
def anti_captcha(browser, message_queue, my_id, my_pw):
    print('안티캡챠 실행')

    # 캡챠 이미지 원소
    elem_captcha_img = browser.find_element_by_xpath('//*[@id="captchaimg"]') 

    # 캡챠 이미지 주소
    captcha_img_url = elem_captcha_img.get_attribute('src') 

    # 안티캡챠 키
    ANTICAPTCHA_KEY = '5330b0f08fe52776ce6caaf56321539d' 

    # 캡챠 정답
    captcha_answer = ImageToTextTask.ImageToTextTask(anticaptcha_key=ANTICAPTCHA_KEY, save_format='const') \
        .captcha_handler(captcha_link=captcha_img_url)

    if not 'solution' in captcha_answer:
        print('로그인 재시도')
        return naver_login(browser, message_queue, my_id, my_pw)
    
    # 캡챠 정답
    captcha_text = captcha_answer['solution']['text'] 

    # 비밀번호 쓰기
    elem_pw = browser.find_element_by_xpath('//*[@id="pw"]') 
    elem_pw.send_keys(my_pw)

    # 캡챠 정답 쓰기
    browser.find_element_by_xpath('//*[@id="chptcha"]').send_keys(captcha_text) 
    elem_pw.submit()

    # 네이버 메인
    succ, _ = find_element_wait_for(browser, '//*[@id="PM_ID_ct"]')

    if succ:
        return True

    else:
        print('로그인 재시도')
        return naver_login(browser, message_queue, my_id, my_pw)        


"""
공통 새글 쓰기 작업
"""
def common_write_post_work(browser, message_queue, my_id, subject, content, content2, content3, tags, imgs, year, month, day, hour, minute, is_open):

    # 본문 프레임
    succ, elem_canvas_frm = find_element_wait_for(browser, '//*[@id="se_canvas_frame"]', timeout=6)

    if not succ:
        message_queue.put('타임아웃: 본문 프레임')
        return False

    # 이미지 올리기 버튼
    succ, elem_img = find_element_wait_for(browser, '//*[@id="se_side_comp_list"]/li[2]/button')

    if not succ:
        message_queue.put('타임아웃: 이미지 올리기 버튼')
        return False       

    # 본문 프레임으로 전환
    browser.switch_to.frame(elem_canvas_frm) 
    time.sleep(2)
    
    # 제목 쓰기
    pyperclip.copy(subject)
    browser.find_element_by_xpath('//textarea').send_keys(Keys.CONTROL, 'v') 
    time.sleep(2)

    # 본문 1, 2 쓰기
    pyperclip.copy('\n'.join([content, content2]))
    succ, elem_content = find_element_wait_for(browser, '//div[@contenteditable="true"]')
    elem_content.send_keys(Keys.CONTROL, 'v') 
    time.sleep(2)

    # 에디터 프레임으로 전환
    browser.switch_to.default_content()
    time.sleep(2)

    # 이미지 업로드
    upload_image(elem_img, imgs)

    # 본문 프레임으로 전환
    browser.switch_to.frame(elem_canvas_frm)     
    time.sleep(2)

    # 본문 3 쓰기
    pyperclip.copy(content3)
    browser.find_element_by_xpath('//div[@contenteditable="true"]').send_keys(Keys.CONTROL, 'v')
    time.sleep(2)

    # 태그 쓰기 #,(컴마)로 구분되어야 함
    if tags is not '':  
        browser.find_element_by_xpath('//li/input[@type="text"]').send_keys(tags) 
        time.sleep(2)

    # 에디터 프레임으로 전환
    browser.switch_to.default_content()

    # 중앙 정렬 버튼 누르기
    wait_and_click(browser, '//a[@class="btn_alignCenter __se_align_btn"]')

    # 상단 발행 버튼 누르기
    wait_and_click(browser, '//a[@id="se_top_publish_btn"]')

    # 발행 버튼
    succ, _ = find_element_wait_for(browser, '//*[@id="se_top_publish_setting_layer"]/div[1]/div[2]/div[3]/button[@class="btn_publish"]')

    if not succ:
        message_queue.put('타임아웃: 발행 버튼')
        return False

    year_month_day = ''
    if year != '예약 안함':
  
        # 예약발행 버튼 클릭
        click(browser, '//*[@id="se_top_publish_setting_layer"]/div[1]/div[2]/div[3]/button[@class="btn_appointment"]')

        # 시간 선택 박스
        succ, _ = find_element_wait_for(browser, '//select[@class="hour"]')

        if not succ:
            message_queue.put('타임아웃: 시간 선택 박스')
            return False

        # 년월일 선택 박스
        elem_year_month_day = browser.find_element_by_xpath('//input[@type="text" and @class="date hasDatepicker" and @readonly="true"]')
        
        # 읽기 전용 속성 제거
        browser.execute_script('arguments[0].removeAttribute("readonly");', elem_year_month_day) 

        # 년월일
        year_month_day = '{0}.{1:02d}.{2:02d}'.format(year, int(month), int(day))

        # 년월일 입력
        browser.execute_script("arguments[0].setAttribute('value', arguments[1])", elem_year_month_day, year_month_day) 

        # 시간 입력
        browser.find_element_by_xpath('//select[@class="hour"]/option[text()="{:02d}"]'.format(int(hour))).click() 

        # 분 입력
        browser.find_element_by_xpath('//select[@class="minutes"]/option[text()="{:02d}"]'.format(minute)).click() 

    if is_open == '공개':

        # 공개로 설정
        browser.find_element_by_xpath('//label[@for="lv_public_1"]').click() 

    else:

        # 비공개로 설정
        browser.find_element_by_xpath('//label[@for="lv_public_4"]').click() 

    # 발행하기
    wait_and_click(browser, '//button[@class="btn_publish"]')
    
    if year == '예약 안함':
        print('{0}, {1}, {2}, {3} 게시 완료'.format(my_id, subject, year, is_open))

    else:
        print('{0}, {1}, {2}, {3}:{4}, {5} 게시 완료'.format(my_id, subject, year_month_day, hour, minute, is_open))

    time.sleep(5)
    return True


"""
새글 쓰기
"""
def write_new_post(browser, message_queue, my_id, subject, content, content2, content3, tags, imgs, year, month, day, hour, minute, is_open):
    
    # 에디터창 열기
    browser.get('https://blog.editor.naver.com/editor') 

    # 팝업 창이 있다면 닫는다
    click_maybe_or_not(browser, '/html/body/div[6]/div/div/div[2]/a')

    # 새글 쓰기
    common_write_post_work(browser, message_queue, my_id, subject, content, content2, content3, tags, imgs, year, month, day, hour, minute, is_open)


def deleter_process(excel_filename, message_queue):
    message_queue.put('미구현')


"""
현재 페이지의 포스트 바꾸기
"""
def modify_other_post_current_page(browser, message_queue, my_id, subject, content, content2, content3, tags, imgs):
    
    # 현재 페이지 포스트 목록 
    elem_links = browser.find_elements_by_xpath('//*[@id="listTopForm"]/table/tbody/tr/td/div/span/a') 
    time.sleep(1)

    for elem_link in elem_links:

        # 나이트 관련 키워드가 없다면
        if not has_keyword(elem_link.text, NIGHT_KEYWORDS):

            # 새 탭으로 열기
            open_new_tab(browser, elem_link)
            time.sleep(1)

            # 수정 탭으로 이동
            browser.find_element_by_tag_name('body').send_keys(Keys.CONTROL + Keys.TAB)
            browser.switch_to.window(browser.window_handles[1])
            time.sleep(1)

            # 수정 버튼 클릭
            succ = wait_and_click(browser, '//*[@id="printPost1"]/tbody/tr/td[2]/div[3]/div[2]/div[2]/a[1]')
            
            if not succ:
                message_queue.put('타임아웃: 수정 버튼')
                return

            # 팝업 창이 있다면 닫는다
            click_maybe_or_not(browser, '/html/body/div[6]/div/div/div[2]/a')

            # 본문 프레임
            succ, elem_canvas_frm = find_element_wait_for(browser, '//*[@id="se_canvas_frame"]', timeout=6)

            if not succ:
                message_queue.put('타임아웃: 본문 프레임')
                return

            # 본문 프레임으로 전환
            browser.switch_to.frame(elem_canvas_frm) 
            time.sleep(2)

            # 제목 지우기
            succ, elem_subject = find_element_wait_for(browser, '//textarea')
            elem_subject.click()
            elem_subject.send_keys(Keys.CONTROL + 'a')
            elem_subject.send_keys(Keys.BACK_SPACE)
            time.sleep(1)

            # 본문 지우기
            pyautogui.press('down')
            for _ in range(10):
                pyautogui.press('enter')

            pyautogui.keyDown('ctrl')
            pyautogui.press('a')
            pyautogui.press('a')
            pyautogui.keyUp('ctrl')
            pyautogui.press('backspace')
            time.sleep(1)

            # 태그 지우기
            succ, elem_tags = find_element_wait_for(browser, '//*[@id="se_canvas_body"]/div[3]/div/div/div/div/span/ul/li')
            succ, elem_tag_fld = find_element_wait_for(browser, '//li/input[@type="text"]')

            try:
                for _ in range(0, len(elem_tags)):
                    elem_tag_fld.send_keys(Keys.BACK_SPACE)
            except:
                pass

            # 에디터 프레임으로 전환
            browser.switch_to.default_content()

            succ = common_write_post_work(browser, message_queue, my_id, subject, content, content2, content3, tags, imgs, '예약 안함', '', '', '', '', '공개')

            # 수정 탭 닫기
            pyautogui.keyDown('ctrl')
            pyautogui.press('w')
            pyautogui.keyUp('ctrl')

            # 블로그 탭으로
            browser.switch_to.window(browser.window_handles[0])


"""
다른 글 바꾸기
"""
def modify_other_post(browser, message_queue, my_id, subject, content, content2, content3, tags, imgs, year, month, day, hour, minute, is_open):

    # 블로그 메인
    browser.get('https://blog.naver.com/PostList.nhn?blogId={}'.format(my_id)) 

    # 전체보기 버튼 클릭
    succ = wait_and_click(browser, '//*[@id="category0"]')

    if not succ:
        message_queue.put('타임아웃: 전체보기 버튼')
        return  

    # 목록 열기 버튼
    succ = wait_and_click(browser, '//*[@id="toplistSpanBlind"]')

    if not succ:
        message_queue.put('타임아웃: 목록 열기 버튼')
        return

    # 페이지 버튼 목록 영역
    succ, _ = find_element_wait_for(browser, '//div[@class="blog2_paginate"]')
    if not succ:
        message_queue.put('타임아웃: 페이지 버튼 목록')
        return               

    current_page_index = 1
    while True:

        # 현재 페이지의 포스트들 바꾸기
        time.sleep(1)
        modify_other_post_current_page(browser, message_queue, my_id, subject, content, content2, content3, tags, imgs)
        time.sleep(1)

        # 다음 페이지의 인덱스
        current_page_index += 1

        # 클릭 가능한 인덱스들
        elem_page_links = browser.find_elements_by_xpath('//div[@id="toplistWrapper"]//div[@class="blog2_paginate"]//a[@href="#"]')

        # 다음 페이지의 인덱스와 같은 링크
        elem_next_page = [elem_page_link for elem_page_link in elem_page_links if elem_page_link.text == str(current_page_index)]

        # 다음 페이지의 인덱스와 같은 링크가 있다면 클릭
        if elem_next_page:
            elem_next_page[0].click()

        # 다음 페이지의 인덱스와 같은 링크가 없다면 '다음' 링크 찾기
        else:
            elem_next_pages = [elem_page_link for elem_page_link in elem_page_links if elem_page_link.text == '다음']

            # '다음' 링크가 있다면 클릭
            if elem_next_pages:
                elem_next_pages[0].click()

            # 아무것도 없다면 끝
            else:
                break


"""
다른 글 바꾸기 프로세스
"""
def modifier_process(excel_filename, message_queue):
    common_login_process_work(excel_filename, message_queue, modify_other_post)
    print('다른 글 바꾸기 프로세스 종료')


if __name__ == '__main__':
    app = QApplication(sys.argv)
    main_window = MainWindow()
    main_window.show()

    sys.exit(app.exec_())
