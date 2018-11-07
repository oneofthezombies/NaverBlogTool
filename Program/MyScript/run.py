"""
네이버 블로그 도구
"""

# 시스템
import sys
import threading

# 크롬 브라우저 제어 모듈
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException

# 캡챠 모듈
from python3_anticaptcha import ImageToTextTask

# 엑셀 모듈
from openpyxl import load_workbook

# 데스크탑 앱 모듈
from PyQt5.QtWidgets import (QApplication, QMainWindow, QFileDialog, 
    QDesktopWidget, QPushButton, QMessageBox)


"""
데스크탑 앱
"""
class MainWindow(QMainWindow):

    def __init__(self):
        super().__init__()

        self.setWindowTitle('네이버 블로그 도구') # 앱 이름        
        self.setGeometry(0, 0, 320, 70) # 앱 크기

        # 앱을 화면 가운데로 옮기기
        rect = self.frameGeometry()
        center = QDesktopWidget().availableGeometry().center()
        rect.moveCenter(center)
        self.move(rect.topLeft())

        open_Excel_btn = QPushButton('엑셀 열기', self) # 엑셀 열기 버튼
        open_Excel_btn.resize(150, 50) # 엑셀 열기 버튼 크기
        open_Excel_btn.move(10, 10) # 엑셀 열기 버튼 위치 이동 (좌상단 기준)

        self.show() # 앱 보이기

    
    def open_excel(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog

        # 엑셀 파일만 열기 허용
        filename, _ = QFileDialog.getOpenFileName(self, 
            '엑셀 파일 열기', './', '엑셀 파일 (*.xlsx)', options=options)

        if not filename:
            return

        Worker(filename).start()
        
        book = load_workbook(filename, read_only=True) # 엑셀 파일
        sheet = book.worksheets[0] # 첫번째 시트

        """
        시트 포맷
           A   B   C        D        E         F         G     H     I     J       K
        1 [id][pw][subject][content][content2][content3][imgs][tags][hour][minute][is_open]
        2 [my_id][my_pw][my_subject]...
        3 [other_id][other_pw][other_subject]...
        4 ...
        """
        rows = sheet.iter_rows(min_row=2) # 첫번째 행을 제외한 나머지 행

        for row in rows:
            my_id = row[0].value
            my_pw = row[1].value
            my_subject = row[2].value
            my_content = row[3].value
            my_content2 = row[4].value
            my_content3 = row[5].value
            my_imgs = row[6].value
            my_tags = row[7].value
            my_hour = row[8].value
            my_minute = row[9].value
            my_is_open = row[10].value


class Worker(threading.Thread):
    def __init__(self, excel_filename):
        print('worker')


if __name__ == '__main__':
    app = QApplication(sys.argv)
    mainWindow = MainWindow()
    sys.exit(app.exec_())
